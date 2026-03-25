#!/usr/bin/env python3
"""
build_workbook.py
-----------------
Fetches role data from an API (or local JSON file) and produces roles_workbook.xlsx with:
  Sheet 1 ("Roles")           — full JSON data as a formatted table
  Sheet 2 ("User Assignments")— data-entry sheet with columns:
                                  email | roleID (dropdown → name) |
                                  roleID_value (VLOOKUP → numeric id) |
                                  agencyid | holdcoid

Usage:
    # Fetch from API (id argument appended as query param)
    python build_workbook.py --id 123
    python build_workbook.py --id 456 --out custom_output.xlsx

    # Override the base API URL
    python build_workbook.py --id 123 --url https://other-api.com/roles

    # Fall back to a local JSON file
    python build_workbook.py --json testJSON.json
"""

import json
import argparse
import sys
from pathlib import Path
from urllib.request import urlopen, Request
from urllib.parse import urlencode, urlparse, parse_qs, urljoin, urlunparse
from urllib.error import URLError, HTTPError

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

DEFAULT_API_URL = "https://json.com/test"

# ---------------------------------------------------------------------------
# Styling helpers
# ---------------------------------------------------------------------------
HEADER_FONT      = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HEADER_FILL      = PatternFill("solid", start_color="2F5496")   # dark blue
ALT_FILL         = PatternFill("solid", start_color="DCE6F1")   # light blue
DATA_FONT        = Font(name="Arial", size=10)
BOLD_FONT        = Font(name="Arial", bold=True, size=10)
CENTER           = Alignment(horizontal="center", vertical="center")
LEFT             = Alignment(horizontal="left",   vertical="center")

THIN = Side(style="thin", color="B0B0B0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(cell, text):
    cell.value     = text
    cell.font      = HEADER_FONT
    cell.fill      = HEADER_FILL
    cell.alignment = CENTER
    cell.border    = BORDER


def style_data(cell, value, align=LEFT, alt_row=False):
    cell.value     = value
    cell.font      = DATA_FONT
    cell.fill      = ALT_FILL if alt_row else PatternFill()
    cell.alignment = align
    cell.border    = BORDER


# ---------------------------------------------------------------------------
# Sheet 1 — Roles reference table
# ---------------------------------------------------------------------------
def build_sheet1(wb, data):
    ws = wb.active
    ws.title = "Roles"
    ws.sheet_view.showGridLines = False

    columns = [
        ("id",          8,  CENTER),
        ("name",        42, LEFT),
        ("type",        10, CENTER),
        ("createdDate", 14, CENTER),
    ]

    # Header row
    ws.row_dimensions[1].height = 20
    for col_idx, (header, width, _) in enumerate(columns, start=1):
        style_header(ws.cell(row=1, column=col_idx), header)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Data rows
    for row_idx, record in enumerate(data, start=2):
        alt = (row_idx % 2 == 0)
        ws.row_dimensions[row_idx].height = 16
        values = [record["id"], record["name"], record["type"], record["createdDate"]]
        for col_idx, (val, (_, _, align)) in enumerate(zip(values, columns), start=1):
            style_data(ws.cell(row=row_idx, column=col_idx), val, align=align, alt_row=alt)

    # Freeze header row
    ws.freeze_panes = "A2"

    return ws


# ---------------------------------------------------------------------------
# Sheet 2 — User Assignments (data-entry)
# ---------------------------------------------------------------------------
def build_sheet2(wb, data, roles_sheet_name="Roles"):
    ws = wb.create_sheet("User Assignments")
    ws.sheet_view.showGridLines = False

    n_roles   = len(data)
    max_rows  = 1000   # validation range depth
    name_col  = "B"    # name column in Roles sheet
    id_col    = "A"    # id column in Roles sheet
    name_range = f"'{roles_sheet_name}'!${name_col}$2:${name_col}${n_roles + 1}"

    # Column layout:
    #  A=email | B=roleID (dropdown of names) | C=roleID_value (VLOOKUP→id) |
    #  D=agencyid | E=holdcoid
    columns = [
        ("email",           30),
        ("roleID",          42),   # dropdown — shows name, human-friendly
        ("roleID_value",    14),   # VLOOKUP resolves to numeric id (auto)
        ("agencyid",        14),
        ("holdcoid",        14),
    ]

    ws.row_dimensions[1].height = 20
    for col_idx, (header, width) in enumerate(columns, start=1):
        style_header(ws.cell(row=1, column=col_idx), header)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # roleID_value column note in header
    note_cell = ws.cell(row=1, column=3)
    note_cell.value = "roleID_value (auto)"

    # Data validation — dropdown on column B (roleID)
    dv = DataValidation(
        type="list",
        formula1=f"={name_range}",
        allow_blank=True,
        showDropDown=False,   # False = show the dropdown arrow
    )
    dv.error      = "Please select a valid role from the list."
    dv.errorTitle = "Invalid Role"
    dv.prompt     = "Pick a role name from the list."
    dv.promptTitle = "Role"
    dv.sqref      = f"B2:B{max_rows}"
    ws.add_data_validation(dv)

    # VLOOKUP formulas — column C auto-resolves the numeric id from column B
    # =IFERROR(VLOOKUP(B2, Roles!$B$2:$A$N, -1, 0), "")  — note: VLOOKUP needs
    # id left of name, so we use INDEX/MATCH instead.
    for row in range(2, max_rows + 1):
        formula = (
            f'=IFERROR(INDEX(\'{roles_sheet_name}\'!${id_col}$2:${id_col}${n_roles+1},'
            f'MATCH(B{row},\'{roles_sheet_name}\'!${name_col}$2:${name_col}${n_roles+1},0)),"")'
        )
        cell = ws.cell(row=row, column=3, value=formula)
        cell.font      = Font(name="Arial", size=10, color="008000")  # green = cross-sheet formula
        cell.alignment = CENTER
        cell.border    = BORDER

    # Style placeholder rows lightly
    for row in range(2, 51):   # pre-style first 50 rows for UX
        alt = (row % 2 == 0)
        for col in [1, 2, 4, 5]:
            c = ws.cell(row=row, column=col)
            c.font      = DATA_FONT
            c.fill      = ALT_FILL if alt else PatternFill()
            c.alignment = LEFT
            c.border    = BORDER

    # Freeze header
    ws.freeze_panes = "A2"

    return ws


# ---------------------------------------------------------------------------
# Data fetching
# ---------------------------------------------------------------------------
def fetch_from_api(base_url, record_id, timeout=15):
    """Append ?id=<record_id> to base_url and return parsed JSON."""
    parsed = urlparse(base_url)
    existing_params = parse_qs(parsed.query)
    existing_params["id"] = [str(record_id)]
    new_query = urlencode({k: v[0] for k, v in existing_params.items()})
    url = urlunparse(parsed._replace(query=new_query))

    print(f"Fetching → {url}")
    req = Request(url, headers={"Accept": "application/json", "User-Agent": "build_workbook/1.0"})
    try:
        with urlopen(req, timeout=timeout) as resp:
            raw = resp.read().decode("utf-8")
            return json.loads(raw)
    except HTTPError as e:
        print(f"HTTP error {e.code}: {e.reason}", file=sys.stderr)
        raise
    except URLError as e:
        print(f"URL error: {e.reason}", file=sys.stderr)
        raise


def load_from_file(json_path):
    path = Path(json_path)
    if not path.exists():
        raise FileNotFoundError(f"JSON file not found: {path}")
    with open(path, "r") as f:
        return json.load(f)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(
        description="Build Excel workbook from API or local JSON roles data.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python build_workbook.py --id 123
  python build_workbook.py --id 456 --out my_roles.xlsx
  python build_workbook.py --id 123 --url https://other-api.com/roles
  python build_workbook.py --json testJSON.json
        """,
    )
    source = parser.add_mutually_exclusive_group(required=True)
    source.add_argument("--id",   type=str, help="Record ID to fetch from the API (appended as ?id=<value>)")
    source.add_argument("--json", type=str, help="Path to a local JSON file (fallback / offline use)")

    parser.add_argument("--url",  default=DEFAULT_API_URL, help=f"Base API URL (default: {DEFAULT_API_URL})")
    parser.add_argument("--out",  default="roles_workbook.xlsx", help="Output Excel file path")
    args = parser.parse_args()

    out_path = Path(args.out)

    if args.id:
        data = fetch_from_api(args.url, args.id)
    else:
        data = load_from_file(args.json)

    if not isinstance(data, list):
        print("Warning: API response is not a list — wrapping in list.", file=sys.stderr)
        data = [data]

    print(f"Loaded {len(data)} roles.")

    wb = Workbook()
    build_sheet1(wb, data)
    build_sheet2(wb, data)

    wb.save(out_path)
    print(f"Workbook saved → {out_path}")


if __name__ == "__main__":
    main()
